import openpyxl
from qradar4py.api import QRadarApi
# Pre configuring arguments that will be used across script in verious functions

book = openpyxl.load_workbook('final.xlsx')
sheet = book.active
response_id_list = []
response_status_list = []
response_description_list = []
response_time_list = []
response_severity_list = []
response_event_count_list = []
response_assigned_to_list = []
tags_to_remove = ["'", " ", "[", "]", "description", "}", "{", "status", ":", '\\n', '\n', 'id', 'start_time',
                  'severity', 'event_count', 'assigned_to']


# Setting API Request Function
def set_api():
    api = QRadarApi("<Qradar-Address>", "<API-Token>", version='10.1', verify=False)
    return api


# Pulling information from Qradar using API
def get_off(api):
    offense_id = ["id"]
    offense_status = ["status"]
    offense_description = ["description"]
    offense_time = ["start_time"]
    offense_severity = ["severity"]
    offense_event_count = ["event_count"]
    offense_assigned_to = ["assigned_to"]
    counter_id = 0
    counter_status = 0
    counter_description = 0
    counter_time = 0
    counter_severity = 0
    counter_event = 0
    counter_assigned = 0

    # Creating a lists for the returned information from Qradar
    for i in offense_id:
        status_code, response_id = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50000',
                                                         fields=offense_id[counter_id])
        response_id_list.append(response_id)
        counter_id += 1

    for i in offense_status:
        status_code, response_status = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50000',
                                                             fields=offense_status[counter_status])
        response_status_list.append(response_status)
        counter_status += 1

    for i in offense_description:
        status_code, response_description = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50000',
                                                                  fields=offense_description[counter_description])
        response_description_list.append(response_description)
        counter_description += 1

    for i in offense_time:
        status_code, response_time = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50000',
                                                           fields=offense_time[counter_time])
        response_time_list.append(response_time)
        counter_time += 1

    for i in offense_severity:
        status_code, response_severity = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50000',
                                                               fields=offense_severity[counter_severity])
        response_severity_list.append(response_severity)
        counter_severity += 1

    for i in offense_event_count:
        status_code, response_event_count = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50000',
                                                                  fields=offense_event_count[counter_event])
        response_event_count_list.append(response_event_count)
        counter_event += 1

    for i in offense_assigned_to:
        status_code, response_assigned_to = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50000',
                                                                  fields=offense_assigned_to[counter_assigned])
        response_assigned_to_list.append(response_assigned_to)
        counter_assigned += 1


# Printing the List to check it returns with correct information
def arrange_response_id_list():
    for i in response_id_list:
        print(i)


# Printing the List to check it returns with correct information
def arrange_response_status_list():
    for i in response_status_list:
        print(i)


# Printing the List to check it returns with correct information
def arrange_response_description_list():
    for i in response_description_list:
        print(i)


def arrange_response_time_list():
    for i in response_time_list:
        print(i)


def arrange_response_severity_list():
    for i in response_severity_list:
        print(i)


def arrange_response_event_count_list():
    for i in response_event_count_list:
        print(i)


def arrange_response_assigned_to_list():
    for i in response_assigned_to_list:
        print(i)


# Function that returns the free cell location in A column
def return_free_cell_location():
    cell_counter = 1
    test = ""
    a = ""
    while str(test) != "None":
        a = sheet.cell(row=cell_counter, column=1)
        test = str(a.value)
        cell_counter += 1
    final = str(a.column_letter) + str(a.row)
    return final


def id_validate():
    # List of IDs returned from the API Request
    fixed_id = str(response_id_list[0])
    for tag in tags_to_remove:
        if tag in fixed_id:
            fixed_id = fixed_id.replace(tag, "")
    fixed_id = fixed_id.split(",")

    # List of Status returned from the API Request

    fixed_status = str(response_status_list)
    for tag in tags_to_remove:
        if tag in fixed_status:
            fixed_status = fixed_status.replace(tag, "")
    fixed_status = fixed_status.split(",")
    # print(fixed_status)

    # List of descriptions returned from the API Request

    fixed_des = str(response_description_list)
    for tag in tags_to_remove:
        if tag in fixed_des:
            fixed_des = fixed_des.replace(tag, "")
    fixed_des = fixed_des.split(",")

    # List of time returned from the API Request
    fixed_time = str(response_time_list)
    for tag in tags_to_remove:
        fixed_time = fixed_time.replace(tag, "")
    fixed_time = fixed_time.split(",")

    # List of Severity returned from the API Request
    fixed_severity = str(response_severity_list)
    for tag in tags_to_remove:
        fixed_severity = fixed_severity.replace(tag, "")
    fixed_severity = fixed_severity.split(",")

    fixed_event_count = str(response_event_count_list)
    for tag in tags_to_remove:
        fixed_event_count = fixed_event_count.replace(tag, "")
    fixed_event_count = fixed_event_count.split(",")

    fixed_assigned = str(response_assigned_to_list)
    for tag in tags_to_remove:
        fixed_assigned = fixed_assigned.replace(tag, "")
    fixed_assigned = fixed_assigned.split(",")

    # Pre configuring arguments that will be needed across the function
    lang = len(list(sheet['A']))
    okay_id_list = list()
    okay_status_list = list()
    okay_des_list = list()
    okay_time_list = list()
    okay_severity_list = list()
    okay_event_count = list()
    okay_assigned_list = list()

    list_location = 0

    for i in fixed_id:
        counter = 2
        best_counter = 1
        while counter <= len(list(sheet['C'])):
            a = sheet.cell(row=counter, column=3)
            if i != str(a.value):
                best_counter += 1
            elif i == str(a.value):
                best_counter -= 1
            if best_counter == lang:
                okay_id_list.append(i)
                okay_status_list.append(fixed_status[list_location])
                okay_des_list.append(fixed_des[list_location])
                okay_time_list.append((fixed_time[list_location]))
                okay_severity_list.append(fixed_severity[list_location])
                okay_event_count.append(fixed_event_count[list_location])
                okay_assigned_list.append(fixed_assigned[list_location])
            counter += 1

        list_location += 1
    # returns 3 lists that are valideted and needed to be written into the Excel file
    print(okay_des_list, okay_status_list, okay_id_list, okay_time_list, okay_severity_list, okay_event_count,
          okay_assigned_list)
    return okay_des_list, okay_status_list, okay_id_list, okay_time_list, okay_severity_list, okay_event_count, okay_assigned_list


# Location for ID
def free_cell_id():
    counter = 2
    string = "None"
    lang = len(list(sheet['C']))
    row = 1
    for i in range(lang):
        a = sheet.cell(row=counter, column=3)
        if string == str(a.value):
            row = a.row
            # print(a.value, row)
            break
        counter += 1
    return row


# Location for Description

def free_cell_des():
    counter = 2
    string = "None"
    row = 1
    u = False
    while u == False:
        a = sheet.cell(row=counter, column=4)
        if string == str(a.value):
            row = a.row
            # print(a.value, row)
            u = True
        counter += 1
    return row


def free_cell_status():
    counter = 2
    string = "None"
    row = 1
    u = False
    while u == False:
        a = sheet.cell(row=counter, column=2)
        if string == str(a.value):
            row = a.row
            # print(a.value, row)
            u = True
        counter += 1
    return row


def free_cell_time():
    counter = 1
    string = "None"
    row = 1
    u = False
    while u == False:
        a = sheet.cell(row=counter, column=1)
        if string == str(a.value):
            row = a.row
            u = True
        counter += 1
    return row


def free_cell_severity():
    counter = 1
    string = "None"
    row = 1
    u = False
    while u == False:
        a = sheet.cell(row=counter, column=5)
        if string == str(a.value):
            row = a.row
            u = True
        counter += 1
    return row


def free_cell_event_count():
    counter = 1
    string = "None"
    row = 1
    u = False
    while u == False:
        a = sheet.cell(row=counter, column=6)
        if string == str(a.value):
            row = a.row
            u = True
        counter += 1
    return row


def free_cell_assigned_to():
    counter = 1
    string = "None"
    row = 1
    u = False
    while u == False:
        a = sheet.cell(row=counter, column=7)
        if string == str(a.value):
            row = a.row
            u = True
        counter += 1
    return row


def write2excel_id():
    okay_des_list, okay_status_list, okay_id_list, okay_time_list, okay_severity_list, okay_event_count, okay_assigned_to = id_validate()
    # ID Part
    counter = 0
    row_id = free_cell_id()
    for i in okay_id_list:
        sheet.cell(row=row_id, column=3).value = okay_id_list[counter]
        counter += 1
        row_id = free_cell_id()
    book.save(filename='final.xlsx')

    # Description Part
    row_des = free_cell_des()
    counter2 = 0
    for i in okay_des_list:
        sheet.cell(row=row_des, column=4).value = okay_des_list[counter2]
        counter2 += 1
        row_des = free_cell_des()
    book.save(filename='final.xlsx')

    # Status Part
    row_status = free_cell_status()
    counter3 = 0
    for i in okay_status_list:
        sheet.cell(row=row_status, column=2).value = okay_status_list[counter3]
        counter3 += 1
        row_status = free_cell_status()
    book.save(filename='final.xlsx')

    # Time Part
    row_time = free_cell_time()
    counter4 = 0
    for i in okay_time_list:
        sheet.cell(row=row_time, column=1).value = okay_time_list[counter4]
        counter4 += 1
        row_time = free_cell_time()
    book.save(filename='final.xlsx')

    # Severity Part
    row_severity = free_cell_severity()
    counter5 = 0
    for i in okay_severity_list:
        sheet.cell(row=row_severity, column=5).value = okay_severity_list[counter5]
        counter5 += 1
        row_severity = free_cell_severity()
    book.save(filename='final.xlsx')

    # Event Count Part
    row_event_count = free_cell_event_count()
    counter6 = 0
    for i in okay_event_count:
        sheet.cell(row=row_event_count, column=6).value = okay_event_count[counter6]
        counter6 += 1
        row_event_count = free_cell_event_count()
    book.save(filename='final.xlsx')

    # Assgiend to Part

    row_assigned = free_cell_assigned_to()
    counter7 = 0
    for i in okay_assigned_to:
        sheet.cell(row=row_assigned, column=7).value = okay_assigned_to[counter7]
        counter7 += 1
        row_assigned = free_cell_assigned_to()
    book.save(filename='final.xlsx')

    print(okay_assigned_to)



# Main function that calls all the other functions

def main():
    api = set_api()
    get_off(api)
    arrange_response_id_list()
    arrange_response_status_list()
    arrange_response_description_list()
    arrange_response_time_list()
    arrange_response_severity_list()
    arrange_response_event_count_list()
    arrange_response_assigned_to_list()
    return_free_cell_location()
    write2excel_id()


# Calling main function
main()
