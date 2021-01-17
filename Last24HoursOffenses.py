import openpyxl
from qradar4py.api import QRadarApi
import calendar
import time
import xlsxwriter
from datetime import date
#---------------------------------

namefinal = str(date.today())
add = ".xlsx"
nameit = namefinal+add
workbook = xlsxwriter.Workbook(nameit)
worksheet = workbook.add_worksheet()
workbook.close()
book = openpyxl.load_workbook(nameit)
sheet = book.active
response_id_list = []
response_time_list = []
response_status_list = []
response_description_list = []
response_severity_list = []
response_event_count_list = []
response_assigned_to_list = []
response_closed_by_list = []
locationlist = []
offense_time = ["close_time"]
offense_id = ["id"]
offense_status = ["status"]
offense_description = ["description"]
offense_severity = ["severity"]
offense_event_count = ["event_count"]
offense_assigned_to = ["assigned_to"]
offense_closed_by = ["closing_user"]
tags_to_remove = ["'", " ", "[", "]", "description", "}", "{", "status", ":", '\\n', '\n', 'id', 'close_time',
                  'severity', 'event_count', 'assigned_to', 'closing_user']
tags_to_remove_special = ["'", "[", "]", "description", "}", "{", "status", ":", '\\n', '\n', 'id', 'close_time',
                          'severity', 'event_count', 'assigned_to', 'note_text']
tags_to_remove_closed = ["'", "[", "]", "description", "}", "{", "status", ":", "closing_user", " "]

#-----------------------------------------------------------------------------------------------------------------------





def set_api():
    api = QRadarApi("<QRadar-Address>>", "<API-TOKEN>>", version='10.1', verify=False)
    return api




# --------------------------------------------------------------------------------------------------------------------


def get_off(api):
    counter_id = 0
    counter_status = 0
    counter_description = 0
    counter_time = 0
    counter_severity = 0
    counter_event = 0
    counter_assigned = 0
    counter_closed_by = 0

    for i in offense_closed_by:
        status_code, response_closed_by = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50',
                                                                fields=offense_closed_by[counter_closed_by])
        response_closed_by_list.append(response_closed_by)
        counter_closed_by += 1

    for i in offense_time:
        status_code, response_time = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50',
                                                           fields=offense_time[counter_time])
        response_time_list.append(response_time)
        counter_time += 1

    for i in offense_id:
        status_code, response_id = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50',
                                                         fields=offense_id[counter_id])
        response_id_list.append(response_id)
        counter_id += 1

    for i in offense_status:
        status_code, response_status = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50',
                                                             fields=offense_status[counter_status])
        response_status_list.append(response_status)
        counter_status += 1

    for i in offense_description:
        status_code, response_description = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50',
                                                                  fields=offense_description[counter_description])
        response_description_list.append(response_description)
        counter_description += 1
    for i in offense_severity:
        status_code, response_severity = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50',
                                                               fields=offense_severity[counter_severity])
        response_severity_list.append(response_severity)
        counter_severity += 1

    for i in offense_event_count:
        status_code, response_event_count = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50',
                                                                  fields=offense_event_count[counter_event])
        response_event_count_list.append(response_event_count)
        counter_event += 1

    for i in offense_assigned_to:
        status_code, response_assigned_to = api.siem.get_offenses(filter='status = CLOSED', Range='items=0-50',
                                                                  fields=offense_assigned_to[counter_assigned])
        response_assigned_to_list.append(response_assigned_to)
        counter_assigned += 1


# ---------------------------------------------------------------------------------------------------------------------


def getcorrectlocation():
    # List of descriptions returned from the API Request
    fixed_closed_by = str(response_closed_by_list)
    for tag in tags_to_remove_closed:
        if tag in fixed_closed_by:
            fixed_closed_by = fixed_closed_by.replace(tag, "")
    fixed_closed_by = fixed_closed_by.split(",")

    fixed_des = str(response_description_list)
    for tag in tags_to_remove_special:
        if tag in fixed_des:
            fixed_des = fixed_des.replace(tag, "")
    fixed_des = fixed_des.split(",")

    # List of time returned from the API Request
    fixed_time = str(response_time_list)
    for tag in tags_to_remove:
        fixed_time = fixed_time.replace(tag, "")
    fixed_time = fixed_time.split(",")
    # for i in range(len(fixed_time)):

    # List of Severity returned from the API Request
    fixed_severity = str(response_severity_list)
    for tag in tags_to_remove:
        fixed_severity = fixed_severity.replace(tag, "")
    fixed_severity = fixed_severity.split(",")

    fixed_event_count = str(response_event_count_list)
    for tag in tags_to_remove:
        fixed_event_count = fixed_event_count.replace(tag, "")
    fixed_event_count = fixed_event_count.split(",")

    fixed_assigned = str(response_assigned_to_list)
    for tag in tags_to_remove:
        fixed_assigned = fixed_assigned.replace(tag, "")
    fixed_assigned = fixed_assigned.split(",")

    fixed_id = str(response_id_list)
    for tag in tags_to_remove:
        if tag in fixed_id:
            fixed_id = fixed_id.replace(tag, "")
    fixed_id = fixed_id.split(",")

    fixed_status = str(response_status_list)
    for tag in tags_to_remove:
        if tag in fixed_status:
            fixed_status = fixed_status.replace(tag, "")
    fixed_status = fixed_status.split(",")
    getunixepochtime = calendar.timegm(time.gmtime())
    todaydate = time.strftime('%Y-%m-%d', time.localtime(getunixepochtime))
    count = 0
    for i in fixed_time:
        first_ten = fixed_time[count][:10]
        convert2check = time.strftime('%Y-%m-%d', time.localtime(int(first_ten)))
        if convert2check == todaydate:
            locationlist.append(count)
        count += 1
    okay_id_list = list()
    okay_status_list = list()
    okay_des_list = list()
    okay_time_list = list()
    okay_severity_list = list()
    okay_event_count = list()
    okay_assigned_list = list()
    okay_closed_by_list = list()
    list_location = 0
    for i in range(len(locationlist)):
        okay_id_list.append(fixed_id[list_location])
        okay_status_list.append(fixed_status[list_location])
        okay_des_list.append(fixed_des[list_location])
        okay_time_list.append(fixed_time[list_location])
        okay_severity_list.append(fixed_severity[list_location])
        okay_event_count.append(fixed_event_count[list_location])
        okay_assigned_list.append(fixed_assigned[list_location])
        okay_closed_by_list.append(fixed_closed_by[list_location])
        list_location += 1
    return okay_des_list, okay_status_list, okay_id_list, okay_time_list, okay_severity_list, okay_event_count, okay_assigned_list, okay_closed_by_list


# --------------------------------------------------------------------------------------------------------------------------------


def free_cell_closed_by():
    row = 2
    return row


def free_cell_id():
    counter = 2
    string = "None"
    lang = len(list(sheet['C']))
    row = 1
    for i in range(lang):
        a = sheet.cell(row=counter, column=3)
        if string == str(a.value):
            row = a.row
            # print(a.value, row)
            break
        counter += 1
    return row


def free_cell_des():
    counter = 2
    string = "None"
    row = 1
    u = False
    while not u:
        a = sheet.cell(row=counter, column=4)
        if string == str(a.value):
            row = a.row
            # print(a.value, row)
            u = True
        counter += 1
    return row


def free_cell_status():
    counter = 2
    string = "None"
    row = 1
    u = False
    while not u:
        a = sheet.cell(row=counter, column=2)
        if string == str(a.value):
            row = a.row
            # print(a.value, row)
            u = True
        counter += 1
    return row


def free_cell_time():
    counter = 2
    string = "None"
    row = 1
    u = False
    while not u:
        a = sheet.cell(row=counter, column=1)
        if string == str(a.value):
            row = a.row
            u = True
        counter += 1
    return row


def free_cell_severity():
    counter = 2
    string = "None"
    row = 1
    u = False
    while not u:
        a = sheet.cell(row=counter, column=5)
        if string == str(a.value):
            row = a.row
            u = True
        counter += 1
    return row


def free_cell_event_count():
    counter = 2
    string = "None"
    row = 1
    u = False
    while u == False:
        a = sheet.cell(row=counter, column=6)
        if string == str(a.value):
            row = a.row
            u = True
        counter += 1
    return row


def free_cell_assigned_to():
    row = 2
    return row


def free_cell_note():
    counter = 2
    string = "None"
    row = 1
    u = False
    while not u:
        a = sheet.cell(row=counter, column=7)
        if string == str(a.value):
            row = a.row
            u = True
        counter += 1
    return row


# ----------------------------------------------------------------------------------------------------------------------------------------------------------

def write2excel_id(api):
    commentlist = list()
    # ID Part + Comment
    okay_des_list, okay_status_list, okay_id_list, okay_time_list, okay_severity_list, okay_event_count, okay_assigned_to, okay_closed_by = getcorrectlocation()
    counter = 0
    row_id = free_cell_id()

    for i in range(len(okay_id_list)):
        okay_id_list[i] = int(okay_id_list[i])

    for i in okay_id_list:
        sheet.cell(row=row_id, column=3).value = str(okay_id_list[counter])
        counter += 1
        row_id = free_cell_id()
    book.save(filename=nameit)

    for i in range(len(okay_id_list)):
        offensenum = okay_id_list[i]
        status_code, response = api.siem.get_offenses_notes_by_offense_id(offense_id=offensenum, Range='items = 0-50')
        fixed_comment = str(response)
        for tag in tags_to_remove_special:
            if tag in fixed_comment:
                fixed_comment = fixed_comment.replace(tag, "")
        fixed_comment = fixed_comment.split(",")
        commentlist.append(fixed_comment[0])

    row_note = free_cell_note()
    counter7 = 0
    for i in commentlist:
        sheet.cell(row=row_note, column=7).value = commentlist[counter7]
        counter7 += 1
        row_note = free_cell_note()

    # Description Part
    row_des = free_cell_des()
    counter2 = 0
    for i in okay_des_list:
        sheet.cell(row=row_des, column=4).value = okay_des_list[counter2]
        counter2 += 1
        row_des = free_cell_des()
    book.save(filename=nameit)

    # closed_by
    row_closed_by = free_cell_closed_by()
    counter9 = 0
    for i in okay_closed_by:
        sheet.cell(row=row_closed_by, column=9).value = okay_closed_by[counter9]
        print(okay_closed_by[counter9])
        counter9 += 1
        row_closed_by += 1
    book.save(filename=nameit)

    # Status Part
    row_status = free_cell_status()
    counter3 = 0
    for i in okay_status_list:
        sheet.cell(row=row_status, column=2).value = okay_status_list[counter3]
        counter3 += 1
        row_status = free_cell_status()
    book.save(filename=nameit)

    # Time Part
    row_time = free_cell_time()
    counter4 = 0
    for i in okay_time_list:
        checkthis = int(okay_time_list[counter4][:10])
        final = time.strftime('%Y-%m-%d' " " '%H:%M:%S', time.localtime(checkthis))
        sheet.cell(row=row_time, column=1).value = final
        counter4 += 1
        row_time = free_cell_time()
    book.save(filename=nameit)

    # Severity Part
    row_severity = free_cell_severity()
    counter5 = 0
    for i in okay_severity_list:
        sheet.cell(row=row_severity, column=5).value = okay_severity_list[counter5]
        counter5 += 1
        row_severity = free_cell_severity()
    book.save(filename=nameit)

    # Event Count Part
    row_event_count = free_cell_event_count()
    counter6 = 0
    for i in okay_event_count:
        sheet.cell(row=row_event_count, column=6).value = okay_event_count[counter6]
        counter6 += 1
        row_event_count = free_cell_event_count()
    book.save(filename=nameit)
    sheet.cell(row=1, column=5).value = "Severity"
    sheet.cell(row=1, column=6).value = "Event Count"
    sheet.cell(row=1, column=4).value = "Description"
    sheet.cell(row=1, column=3).value = "Offense ID"
    sheet.cell(row=1, column=2).value = "Status"
    sheet.cell(row=1, column=1).value = "Time"
    sheet.cell(row=1, column=7).value = "Closeing Notes"
    sheet.cell(row=1, column=8).value = "Assigned To"
    sheet.cell(row=1, column=9).value = "Closed By"
    book.save(filename=nameit)

    # Assgiend to Part

    row_assigned = free_cell_assigned_to()
    counter8 = 0
    for i in okay_assigned_to:
        sheet.cell(row=row_assigned, column=8).value = okay_assigned_to[counter8]
        counter8 += 1
        row_assigned += 1
    book.save(filename=nameit)


# ----------------------------------------------------------------------------------------------------------------------------------------------------------

def main():
    api = set_api()
    get_off(api)
    write2excel_id(api)
    
main()
